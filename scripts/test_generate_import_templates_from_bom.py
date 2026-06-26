import pathlib
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))

from generate_import_templates_from_bom import (
    DEFAULT_MONTHS,
    DEFAULT_START_MONTH,
    load_bom_rows_from_workbook,
    write_template_workbooks,
)


class GenerateImportTemplatesFromBomTests(unittest.TestCase):
    def test_load_bom_rows_and_write_template_workbooks(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            bom_path = root / "template-bom.xlsx"
            self._write_bom_workbook(bom_path)

            bom_rows = load_bom_rows_from_workbook(bom_path)
            self.assertEqual(3, len(bom_rows))
            self.assertEqual("FP100", bom_rows[0]["parent_code"])
            self.assertEqual("采购件", bom_rows[0]["part_attribute"])
            self.assertEqual("V1", bom_rows[0]["version"])

            generated = write_template_workbooks(
                bom_rows,
                output_dir=root,
                start_month=DEFAULT_START_MONTH,
                months=DEFAULT_MONTHS,
            )

            expected_names = {
                "template-demand.xlsx",
                "template-inventory.xlsx",
                "template-safety-stock.xlsx",
                "template-operating-days.xlsx",
                "template-equipment-catalog.xlsx",
                "template-part-master.xlsx",
                "template-shared-mold-rules.xlsx",
            }
            self.assertEqual(expected_names, {path.name for path in generated})

            demand = load_workbook(root / "template-demand.xlsx", data_only=True)
            self.assertEqual(["完成品入库需求数"], demand.sheetnames)
            sheet = demand["完成品入库需求数"]
            self.assertEqual("客户", sheet["A2"].value)
            self.assertEqual("FP100", sheet["B3"].value)
            self.assertEqual(202606, sheet["C3"].value)

            part_master = load_workbook(root / "template-part-master.xlsx", data_only=True)
            self.assertEqual(["零件主数据"], part_master.sheetnames)
            self.assertEqual("零件编码", part_master["零件主数据"]["A2"].value)

            shared_mold = load_workbook(root / "template-shared-mold-rules.xlsx", data_only=True)
            self.assertEqual(["共模规则"], shared_mold.sheetnames)
            self.assertEqual("产品A编码", shared_mold["共模规则"]["A2"].value)

    def _write_bom_workbook(self, path: Path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "BOM"
        sheet.append(["说明：测试", None, None, None, None, None, None, None, None, None, None, None, None, None])
        sheet.append(["父零件", "子零件", "用量", "工序", "设备", "制造部门", "制造单元", "模腔数/取数（pcs）", "制造周期（S）", "持台人数（人）", "单件节拍（S）", "报废率", "版本号", "子零件属性"])
        sheet.append(["FP100", "FP100-1", 1, "焊接", "EQ-01", "制造一课", "焊接", 1, 64, 1, 64, 0.01, "V1", "采购件"])
        sheet.append(["FP100-1", "SEMI200", 1, "整理", None, "制造二课", "组立", 1, 60, 1, 60, 0, "V1", "自制半成品"])
        sheet.append(["FP200", "SEMI300", 1, "注塑", "EQ-02", "制造三课", "成型", 2, 80, 1, 40, 0.02, "V1", ""])
        workbook.save(path)


if __name__ == "__main__":
    unittest.main()
