from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


DEFAULT_START_MONTH = 202606
DEFAULT_MONTHS = 3
DEFAULT_VERSION = "20260331-1"
DEFAULT_CUSTOMER = "测试客户"
DEFAULT_BRAND = "AUTO"


@dataclass(frozen=True)
class BomRow:
    parent_code: str
    child_code: str | None
    usage_qty: float | None
    process: str | None
    equipment: str | None
    manufacturing_department: str | None
    manufacturing_unit: str | None
    mold_cavity: int | None
    cycle_time: float | None
    staff_count: float | None
    takt_time: float | None
    scrap_rate: float | None
    part_attribute: str | None
    version: str


def load_bom_rows_from_workbook(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["BOM"]
    rows: list[dict[str, object]] = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if row is None or not row[0]:
            continue
        rows.append(
            {
                "parent_code": str(row[0]).strip(),
                "child_code": normalize_text(row[1]),
                "usage_qty": to_float(row[2]),
                "process": normalize_text(row[3]),
                "equipment": normalize_text(row[4]),
                "manufacturing_department": normalize_text(row[5]),
                "manufacturing_unit": normalize_text(row[6]),
                "mold_cavity": to_int(row[7]),
                "cycle_time": to_float(row[8]),
                "staff_count": to_float(row[9]),
                "takt_time": to_float(row[10]),
                "scrap_rate": to_float(row[11]),
                "version": normalize_text(row[12]) or DEFAULT_VERSION,
                "part_attribute": normalize_text(row[13]),
            }
        )
    return rows


def write_template_workbooks(
    bom_rows: list[dict[str, object]],
    *,
    output_dir: Path,
    start_month: int = DEFAULT_START_MONTH,
    months: int = DEFAULT_MONTHS,
) -> list[Path]:
    month_values = build_months(start_month, months)
    bundle = build_bundle(bom_rows, month_values)
    output_dir.mkdir(parents=True, exist_ok=True)

    outputs = [
        write_demand_workbook(output_dir / "template-demand.xlsx", bundle["demand_rows"]),
        write_inventory_workbook(output_dir / "template-inventory.xlsx", bundle["inventory_rows"]),
        write_safety_workbook(output_dir / "template-safety-stock.xlsx", bundle["safety_rows"]),
        write_operating_days_workbook(output_dir / "template-operating-days.xlsx", bundle["operating_days_rows"]),
        write_equipment_workbook(output_dir / "template-equipment-catalog.xlsx", bundle["equipment_rows"]),
        write_part_master_workbook(output_dir / "template-part-master.xlsx", bundle["part_rows"]),
        write_shared_mold_workbook(output_dir / "template-shared-mold-rules.xlsx", bundle["shared_mold_rows"]),
    ]
    return outputs


def build_bundle(bom_rows: list[dict[str, object]], months: list[int]) -> dict[str, list[dict[str, object]]]:
    version = str(bom_rows[0].get("version") or DEFAULT_VERSION) if bom_rows else DEFAULT_VERSION
    parent_codes = {str(row["parent_code"]) for row in bom_rows if row.get("parent_code")}
    child_codes = {str(row["child_code"]) for row in bom_rows if row.get("child_code")}
    root_codes = sorted(parent_codes - child_codes)
    intermediate_codes = sorted(parent_codes & child_codes)
    part_codes = sorted(parent_codes | child_codes)
    inventory_month = shift_month(months[0], -1)

    demand_rows: list[dict[str, object]] = []
    for root_index, code in enumerate(root_codes, start=1):
        base_qty = 120 + root_index * 8
        for month_index, month in enumerate(months):
            demand_qty = float(base_qty + month_index * 12)
            ending_inventory = round(demand_qty * 0.12, 2)
            min_safety_stock = round(demand_qty * 0.08, 2)
            net_demand = round(demand_qty - ending_inventory, 2)
            demand_rows.append(
                {
                    "customer": DEFAULT_CUSTOMER,
                    "item_code": code,
                    "year_month": month,
                    "demand_qty": demand_qty,
                    "ending_inventory": ending_inventory,
                    "min_safety_stock": min_safety_stock,
                    "net_demand": net_demand,
                    "version": version,
                }
            )

    non_root_materials = sorted(code for code in intermediate_codes if not code.endswith("-1"))

    inventory_rows = []
    safety_rows = []
    for index, code in enumerate(non_root_materials, start=1):
        available_qty = round((18 + index * 1.5) * 4.5, 2)
        inventory_rows.append(
            {
                "item_code": code,
                "year_month": inventory_month,
                "available_qty": available_qty,
                "version": version,
            }
        )
        daily_equivalent = round(14 + index * 0.75, 2)
        safety_days = float(3 + (index % 3))
        max_days = safety_days + 4.0
        for month in months:
            safety_rows.append(
                {
                    "item_code": code,
                    "year_month": month,
                    "daily_equivalent": daily_equivalent,
                    "safety_days": safety_days,
                    "max_days": max_days,
                    "version": version,
                }
            )

    operating_days_rows = [
        {
            "year_month": month,
            "total_days": 26.0,
            "work_days": 21.0,
            "weekend_days": 5.0,
            "holiday_days": 0.0,
        }
        for month in months
    ]

    grouped_equipment: dict[tuple[str, str], dict[str, object]] = {}
    grouped_count: dict[tuple[str, str], int] = {}
    for row in bom_rows:
        department = normalize_text(row.get("manufacturing_department"))
        equipment = normalize_text(row.get("equipment"))
        if not department or not equipment:
            continue
        key = (department, equipment)
        grouped_count[key] = grouped_count.get(key, 0) + 1
        grouped_equipment.setdefault(
            key,
            {
                "manufacturing_department": department,
                "equipment_category": infer_equipment_category(equipment, row.get("process")),
                "equipment_brand": DEFAULT_BRAND,
                "equipment_model": equipment,
            },
        )
    equipment_rows = []
    for key in sorted(grouped_equipment.keys()):
        base = grouped_equipment[key]
        equipment_rows.append(
            {
                **base,
                "equipment_count": max(1, grouped_count[key]),
            }
        )

    part_rows = []
    for code in part_codes:
        if code in root_codes:
            name = f"完成品-{code}"
        elif code.endswith("-1"):
            name = f"虚拟总成-{code}"
        elif code in intermediate_codes:
            name = f"半成品-{code}"
        else:
            name = f"零件-{code}"
        part_rows.append(
            {
                "part_no": code,
                "product_name": name,
                "product_no": f"P-{code}",
                "project_name": f"项目-{code[:4]}",
            }
        )

    shared_mold_rows = []
    for index in range(0, len(root_codes) - 1, 2):
        shared_mold_rows.append(
            {
                "product_a_code": root_codes[index],
                "product_b_code": root_codes[index + 1],
                "equipment_code": "",
                "mold_code": "",
                "enabled": "Y",
                "remark": "自动生成测试规则",
            }
        )

    return {
        "demand_rows": demand_rows,
        "inventory_rows": inventory_rows,
        "safety_rows": safety_rows,
        "operating_days_rows": operating_days_rows,
        "equipment_rows": equipment_rows,
        "part_rows": part_rows,
        "shared_mold_rows": shared_mold_rows,
    }


def write_demand_workbook(path: Path, rows: list[dict[str, object]]) -> Path:
    return write_workbook(
        path,
        "完成品入库需求数",
        "说明：客户+版本号全删全导；列顺序需保持不变",
        ["客户", "存货编码", "年月", "需求数量", "期末库存", "最小安全库存", "完成品入库需求数", "版本号"],
        rows,
        ["customer", "item_code", "year_month", "demand_qty", "ending_inventory", "min_safety_stock", "net_demand", "version"],
    )


def write_inventory_workbook(path: Path, rows: list[dict[str, object]]) -> Path:
    return write_workbook(
        path,
        "半成品期末盘点数",
        "说明：按版本号全删全导；年月使用需求起始月的前一月",
        ["存货编码", "年月（底）", "可用量", "版本号"],
        rows,
        ["item_code", "year_month", "available_qty", "version"],
    )


def write_safety_workbook(path: Path, rows: list[dict[str, object]]) -> Path:
    return write_workbook(
        path,
        "半成品安全库存",
        "说明：按版本号全删全导；仅为半成品生成",
        ["存货编码", "年月", "日当量", "安全天数", "最大天数", "版本号"],
        rows,
        ["item_code", "year_month", "daily_equivalent", "safety_days", "max_days", "version"],
    )


def write_operating_days_workbook(path: Path, rows: list[dict[str, object]]) -> Path:
    return write_workbook(
        path,
        "稼动天数",
        "说明：按年月 upsert；当前模板固定生成 202606-202608",
        ["年月", "总出勤天数", "工作日", "双休日天数", "国定节假日天数"],
        rows,
        ["year_month", "total_days", "work_days", "weekend_days", "holiday_days"],
    )


def write_equipment_workbook(path: Path, rows: list[dict[str, object]]) -> Path:
    return write_workbook(
        path,
        "设备清单",
        "说明：按制造部门全删全导；制造部门+设备小类唯一；台数必须大于0",
        ["制造部门", "设备大类", "设备品牌", "设备小类", "台数"],
        rows,
        ["manufacturing_department", "equipment_category", "equipment_brand", "equipment_model", "equipment_count"],
    )


def write_part_master_workbook(path: Path, rows: list[dict[str, object]]) -> Path:
    return write_workbook(
        path,
        "零件主数据",
        "说明：partNo 唯一；导入按 partNo 执行新增或更新",
        ["零件编码", "零件名称", "零件番号", "项目名称"],
        rows,
        ["part_no", "product_name", "product_no", "project_name"],
    )


def write_shared_mold_workbook(path: Path, rows: list[dict[str, object]]) -> Path:
    return write_workbook(
        path,
        "共模规则",
        "说明：产品A编码+产品B编码唯一；设备编号/模具编号可为空；是否启用填写 Y 或 N",
        ["产品A编码", "产品B编码", "设备编号", "模具编号", "是否启用", "备注"],
        rows,
        ["product_a_code", "product_b_code", "equipment_code", "mold_code", "enabled", "remark"],
    )


def write_workbook(
    path: Path,
    sheet_name: str,
    notice: str,
    headers: list[str],
    rows: list[dict[str, object]],
    keys: list[str],
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append([notice])
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(key) for key in keys])
    workbook.save(path)
    return path


def build_months(start_month: int, count: int) -> list[int]:
    return [shift_month(start_month, offset) for offset in range(count)]


def shift_month(start_yyyymm: int, offset: int) -> int:
    year = start_yyyymm // 100
    month = start_yyyymm % 100
    month_index = (year * 12 + (month - 1)) + offset
    return (month_index // 12) * 100 + (month_index % 12) + 1


def infer_equipment_category(equipment: str, process: object) -> str:
    process_text = normalize_text(process)
    if process_text:
        return f"{process_text}设备"
    return f"{equipment.split('-')[0]}设备"


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.upper() == "NULL":
        return None
    return text


def to_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    return float(value)


def to_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    return int(float(value))


if __name__ == "__main__":
    root = Path(__file__).resolve().parent.parent
    bom_rows = load_bom_rows_from_workbook(root / "template-bom.xlsx")
    write_template_workbooks(bom_rows, output_dir=root)
